"""Privacy-safe adapter for the synthetic skill-passport workbook.

The adapter is deliberately outside the scoring service.  It knows the final
workbook's column names, selects only employment-safe fields, and produces the
stable :class:`~app.schemas.CandidateProfile` model consumed by the recommender.
No name, gender, prison identifier, conduct, savings, health/behaviour label, or
verification hash is copied into a recommendation payload.
"""

from __future__ import annotations

import argparse
from collections import Counter
from collections.abc import Mapping, Sequence
from dataclasses import asdict, dataclass
from datetime import date, datetime
import json
from pathlib import Path
import sys
from typing import Any

from openpyxl import load_workbook

from app.schemas import CandidateProfile
from app.services.normalizer import normalize_text
from app.services.skill_extractor import extract_skills


REQUIRED_COLUMNS = frozenset(
    {
        "passport_id",
        "primary_trade_vocational",
        "course_completion_status",
        "specific_machinery_skills",
        "target_job_domain",
        "preferred_work_districts",
        "consent_to_share_profile",
    }
)

# These are the only workbook columns allowed to influence CandidateProfile.
# Optional fields make the adapter forward-compatible when the dataset adds the
# three employment attributes that are currently absent.
SAFE_SOURCE_COLUMNS = REQUIRED_COLUMNS | frozenset(
    {
        "canonical_certificates",
        "experience_months",
        "available_from",
    }
)

_DISTRICT_NAMES = {
    "mumbai": "Mumbai",
    "mumbai city": "Mumbai",
    "navi mumbai": "Navi Mumbai",
    "nashik": "Nashik",
    "pune": "Pune",
    "raigad": "Raigad",
    "thane": "Thane",
}


class SkillPassportMappingError(ValueError):
    """Raised when the workbook cannot be mapped without guessing."""


@dataclass(frozen=True, slots=True)
class MappingIssue:
    """A non-fatal, row-level limitation found during mapping."""

    row_number: int
    candidate_id: str | None
    code: str
    message: str


@dataclass(slots=True)
class SkillPassportImportResult:
    """Mapped candidates plus auditable warnings about missing source data."""

    candidates: list[CandidateProfile]
    issues: list[MappingIssue]
    source_rows: int

    def summary(self) -> dict[str, Any]:
        issue_counts = Counter(issue.code for issue in self.issues)
        return {
            "source_rows": self.source_rows,
            "mapped_candidates": len(self.candidates),
            "consenting_candidates": sum(
                candidate.consent for candidate in self.candidates
            ),
            "candidates_with_verified_skills": sum(
                bool(candidate.verified_skills) for candidate in self.candidates
            ),
            "issue_counts": dict(sorted(issue_counts.items())),
        }


def _clean_header(value: object) -> str:
    return normalize_text(str(value or "")).replace(" ", "_")


def _clean_required_text(value: object, column: str, row_number: int) -> str:
    text = " ".join(str(value or "").split())
    if not text:
        raise SkillPassportMappingError(
            f"Row {row_number}: required column '{column}' is blank"
        )
    return text


def _split_values(value: object) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        raw_values = [str(item) for item in value]
    else:
        raw_values = str(value).split("|")

    values: list[str] = []
    seen: set[str] = set()
    for raw_value in raw_values:
        cleaned = " ".join(raw_value.split())
        key = cleaned.casefold()
        if cleaned and key not in seen:
            seen.add(key)
            values.append(cleaned)
    return values


def _parse_consent(value: object, row_number: int) -> bool:
    if type(value) is not bool:
        raise SkillPassportMappingError(
            f"Row {row_number}: consent_to_share_profile must be a true/false "
            "Excel Boolean"
        )
    return value


def _parse_experience(value: object, row_number: int) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, bool):
        raise SkillPassportMappingError(
            f"Row {row_number}: experience_months must be a non-negative integer"
        )
    try:
        if isinstance(value, int):
            number = value
        elif isinstance(value, float):
            if not value.is_integer():
                raise ValueError
            number = int(value)
        else:
            text = str(value).strip()
            if text.startswith("+"):
                text = text[1:]
            if not text.isdigit():
                raise ValueError
            number = int(text)
    except (TypeError, ValueError, OverflowError) as exc:
        raise SkillPassportMappingError(
            f"Row {row_number}: experience_months must be a non-negative integer"
        ) from exc
    if number < 0:
        raise SkillPassportMappingError(
            f"Row {row_number}: experience_months must be a non-negative integer"
        )
    return number


def _parse_available_from(value: object, row_number: int) -> date | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError as exc:
        raise SkillPassportMappingError(
            f"Row {row_number}: available_from must use YYYY-MM-DD"
        ) from exc


def _category_tag(value: object) -> str:
    return normalize_text(str(value or "")).replace(" ", "_")


def _districts(value: object) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for district in _split_values(value):
        key = normalize_text(district)
        display = _DISTRICT_NAMES.get(key, " ".join(district.split()))
        display_key = display.casefold()
        if display and display_key not in seen:
            seen.add(display_key)
            normalized.append(display)
    return normalized


def _verified_skills(
    row: Mapping[str, object],
    row_number: int,
    candidate_id: str,
) -> tuple[list[str], list[MappingIssue]]:
    status = normalize_text(str(row.get("course_completion_status") or ""))
    if status not in {"certified", "in training"}:
        raise SkillPassportMappingError(
            f"Row {row_number}: course_completion_status must be "
            "'Certified' or 'In_Training'"
        )
    primary_trade = _clean_required_text(
        row.get("primary_trade_vocational"),
        "primary_trade_vocational",
        row_number,
    )
    source_phrases = [
        primary_trade,
        *_split_values(row.get("specific_machinery_skills")),
    ]

    issues: list[MappingIssue] = []
    recognized: list[str] = []
    seen: set[str] = set()
    for phrase in source_phrases:
        matches = extract_skills(phrase)
        if not matches:
            issues.append(
                MappingIssue(
                    row_number=row_number,
                    candidate_id=candidate_id,
                    code="unmapped_skill_phrase",
                    message=f"No canonical skill mapping for: {phrase}",
                )
            )
        for match in matches:
            skill = match["canonical_skill"]
            if skill not in seen:
                seen.add(skill)
                recognized.append(skill)

    if status != "certified":
        issues.append(
            MappingIssue(
                row_number=row_number,
                candidate_id=candidate_id,
                code="training_not_certified",
                message=(
                    "Trade and machinery skills were not added to verified_skills "
                    f"because course status is '{row.get('course_completion_status')}'."
                ),
            )
        )
        return [], issues
    return recognized, issues


def map_skill_passport_row(
    raw_row: Mapping[str, object],
    *,
    row_number: int = 2,
) -> tuple[CandidateProfile, list[MappingIssue]]:
    """Map one workbook row to CandidateProfile using only whitelisted fields."""

    row = {_clean_header(key): value for key, value in raw_row.items()}
    missing_columns = sorted(REQUIRED_COLUMNS - row.keys())
    if missing_columns:
        raise SkillPassportMappingError(
            "Missing required skill-passport columns: " + ", ".join(missing_columns)
        )

    candidate_id = _clean_required_text(
        row.get("passport_id"), "passport_id", row_number
    )
    verified_skills, issues = _verified_skills(row, row_number, candidate_id)

    experience = _parse_experience(row.get("experience_months"), row_number)
    if experience is None:
        experience = 0
        issues.append(
            MappingIssue(
                row_number=row_number,
                candidate_id=candidate_id,
                code="experience_months_missing",
                message=(
                    "No employment experience_months column/value is available; "
                    "the candidate is mapped with 0 months. Workshop training hours "
                    "are intentionally not converted to employment experience."
                ),
            )
        )

    is_certified = normalize_text(
        str(row.get("course_completion_status") or "")
    ) == "certified"
    certificates = (
        _split_values(row.get("canonical_certificates"))
        if is_certified
        else []
    )
    if not certificates and is_certified:
        issues.append(
            MappingIssue(
                row_number=row_number,
                candidate_id=candidate_id,
                code="canonical_certificates_missing",
                message=(
                    "The workbook has an issuer and NSQF level but no canonical "
                    "certificate name/code, so certificates is empty."
                ),
            )
        )

    category_source = _clean_required_text(
        row.get("target_job_domain"), "target_job_domain", row_number
    )
    category = _category_tag(category_source)
    preferred_categories = [category]
    district_source = _clean_required_text(
        row.get("preferred_work_districts"),
        "preferred_work_districts",
        row_number,
    )
    preferred_districts = _districts(district_source)

    candidate = CandidateProfile(
        candidate_id=candidate_id,
        verified_skills=verified_skills,
        certificates=certificates,
        experience_months=experience,
        preferred_job_categories=preferred_categories,
        preferred_districts=preferred_districts,
        available_from=_parse_available_from(row.get("available_from"), row_number),
        consent=_parse_consent(row.get("consent_to_share_profile"), row_number),
    )
    return candidate, issues


def load_skill_passport_workbook(
    workbook_path: str | Path,
    *,
    sheet_name: str | None = None,
) -> SkillPassportImportResult:
    """Read the synthetic workbook and map all non-blank rows safely."""

    path = Path(workbook_path)
    if not path.is_file():
        raise SkillPassportMappingError(f"Workbook not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name is None:
            worksheet = workbook[workbook.sheetnames[0]]
        elif sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            raise SkillPassportMappingError(
                f"Worksheet '{sheet_name}' not found. Available sheets: "
                + ", ".join(workbook.sheetnames)
            )

        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise SkillPassportMappingError("Workbook is empty") from exc

        headers = [_clean_header(header) for header in raw_headers]
        if not any(headers):
            raise SkillPassportMappingError("Workbook header row is blank")
        if len(set(headers)) != len(headers):
            raise SkillPassportMappingError(
                "Workbook contains duplicate column names after normalization"
            )
        missing_columns = sorted(REQUIRED_COLUMNS - set(headers))
        if missing_columns:
            raise SkillPassportMappingError(
                "Missing required skill-passport columns: "
                + ", ".join(missing_columns)
            )

        candidates: list[CandidateProfile] = []
        issues: list[MappingIssue] = []
        candidate_ids: set[str] = set()
        source_rows = 0

        for row_number, values in enumerate(rows, start=2):
            if not any(value is not None and str(value).strip() for value in values):
                continue
            source_rows += 1
            raw_row = dict(zip(headers, values))
            candidate, row_issues = map_skill_passport_row(
                raw_row, row_number=row_number
            )
            if candidate.candidate_id in candidate_ids:
                raise SkillPassportMappingError(
                    f"Row {row_number}: duplicate passport_id "
                    f"'{candidate.candidate_id}'"
                )
            candidate_ids.add(candidate.candidate_id)
            candidates.append(candidate)
            issues.extend(row_issues)

        return SkillPassportImportResult(
            candidates=candidates,
            issues=issues,
            source_rows=source_rows,
        )
    finally:
        workbook.close()


def _write_candidates(
    output_path: Path, candidates: Sequence[CandidateProfile]
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = [candidate.model_dump(mode="json") for candidate in candidates]
    output_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def main(argv: Sequence[str] | None = None) -> int:
    """Validate/map a workbook without starting the FastAPI server."""

    parser = argparse.ArgumentParser(
        description="Validate and safely map the synthetic skill-passport workbook."
    )
    parser.add_argument("workbook", type=Path, help="Path to the .xlsx workbook")
    parser.add_argument("--sheet", help="Worksheet name; defaults to the first sheet")
    parser.add_argument(
        "--output",
        type=Path,
        help="Optional path for sanitized CandidateProfile JSON",
    )
    args = parser.parse_args(argv)

    try:
        result = load_skill_passport_workbook(args.workbook, sheet_name=args.sheet)
    except SkillPassportMappingError as exc:
        print(f"Dataset validation failed: {exc}", file=sys.stderr)
        return 2

    if args.output:
        _write_candidates(args.output, result.candidates)

    report = {
        **result.summary(),
        "output": str(args.output.resolve()) if args.output else None,
    }
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
