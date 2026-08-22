"""Adapter for the RIHAI SETU synthetic job-catalog workbook."""

from __future__ import annotations

import argparse
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
import json
from pathlib import Path
import sys
from typing import Any

from openpyxl import load_workbook
from pydantic import ValidationError

from app.schemas import Job
from app.services.normalizer import normalize_text
from app.services.skill_extractor import extract_skills


REQUIRED_COLUMNS = frozenset(
    {
        "job_id",
        "title",
        "description",
        "required_skills",
        "preferred_skills",
        "required_certificates",
        "minimum_experience_months",
        "job_category",
        "district",
        "status",
    }
)


class JobCatalogMappingError(ValueError):
    """Raised when a job workbook cannot be mapped without guessing."""


@dataclass(slots=True)
class JobCatalogImportResult:
    jobs: list[Job]
    source_rows: int
    skill_mappings: dict[str, str]

    def summary(self) -> dict[str, Any]:
        return {
            "source_rows": self.source_rows,
            "mapped_jobs": len(self.jobs),
            "active_jobs": sum(job.status.value == "active" for job in self.jobs),
            "paused_jobs": sum(job.status.value == "paused" for job in self.jobs),
            "closed_jobs": sum(job.status.value == "closed" for job in self.jobs),
            "source_skill_labels": len(self.skill_mappings),
            "canonical_skills": len(set(self.skill_mappings.values())),
        }


def _clean_header(value: object) -> str:
    return normalize_text(str(value or "")).replace(" ", "_")


def _clean_text(value: object) -> str:
    return " ".join(str(value or "").split())


def _required_text(value: object, column: str, row_number: int) -> str:
    text = _clean_text(value)
    if not text:
        raise JobCatalogMappingError(
            f"Row {row_number}: required column '{column}' is blank"
        )
    return text


def _split_values(value: object) -> list[str]:
    values: list[str] = []
    seen: set[str] = set()
    for raw in str(value or "").split("|"):
        cleaned = _clean_text(raw)
        key = normalize_text(cleaned)
        if cleaned and key not in seen:
            seen.add(key)
            values.append(cleaned)
    return values


def canonicalize_skill_label(value: object, *, context: str = "skill") -> str:
    """Map one workbook tag/alias to exactly one canonical service skill."""

    source = _clean_text(value)
    if not source:
        raise JobCatalogMappingError(f"{context}: skill label is blank")
    matches = extract_skills(source.replace("_", " "))
    canonical = list(dict.fromkeys(match["canonical_skill"] for match in matches))
    if not canonical:
        raise JobCatalogMappingError(
            f"{context}: no canonical vocabulary mapping for '{source}'"
        )
    if len(canonical) != 1:
        raise JobCatalogMappingError(
            f"{context}: ambiguous skill label '{source}' maps to "
            + ", ".join(canonical)
        )
    return canonical[0]


def _canonical_skill_list(
    value: object,
    *,
    row_number: int,
    column: str,
    mappings: dict[str, str],
) -> list[str]:
    canonical: list[str] = []
    seen: set[str] = set()
    for source in _split_values(value):
        mapped = canonicalize_skill_label(
            source, context=f"Row {row_number} column '{column}'"
        )
        mappings[source] = mapped
        if mapped not in seen:
            seen.add(mapped)
            canonical.append(mapped)
    return canonical


def _experience_months(value: object, row_number: int) -> int:
    if value is None or str(value).strip() == "":
        return 0
    if isinstance(value, bool):
        raise JobCatalogMappingError(
            f"Row {row_number}: minimum_experience_months must be a "
            "non-negative integer"
        )
    try:
        number = int(value)
    except (TypeError, ValueError, OverflowError) as exc:
        raise JobCatalogMappingError(
            f"Row {row_number}: minimum_experience_months must be a "
            "non-negative integer"
        ) from exc
    if number < 0 or (isinstance(value, float) and not value.is_integer()):
        raise JobCatalogMappingError(
            f"Row {row_number}: minimum_experience_months must be a "
            "non-negative integer"
        )
    return number


def _category_tag(value: object, row_number: int) -> str:
    category = _required_text(value, "job_category", row_number)
    return normalize_text(category).replace(" ", "_")


def map_job_catalog_row(
    raw_row: Mapping[str, object],
    *,
    row_number: int = 2,
    skill_mappings: dict[str, str] | None = None,
) -> Job:
    """Map one job row into the stable API model."""

    row = {_clean_header(key): value for key, value in raw_row.items()}
    missing = sorted(REQUIRED_COLUMNS - row.keys())
    if missing:
        raise JobCatalogMappingError(
            "Missing required job-catalog columns: " + ", ".join(missing)
        )
    mappings = skill_mappings if skill_mappings is not None else {}
    required_skills = _canonical_skill_list(
        row.get("required_skills"),
        row_number=row_number,
        column="required_skills",
        mappings=mappings,
    )
    if not required_skills:
        raise JobCatalogMappingError(
            f"Row {row_number}: required_skills must contain at least one skill"
        )
    preferred_skills = _canonical_skill_list(
        row.get("preferred_skills"),
        row_number=row_number,
        column="preferred_skills",
        mappings=mappings,
    )

    try:
        return Job(
            job_id=_required_text(row.get("job_id"), "job_id", row_number),
            title=_required_text(row.get("title"), "title", row_number),
            description=_clean_text(row.get("description")),
            required_skills=required_skills,
            preferred_skills=preferred_skills,
            required_certificates=_split_values(
                row.get("required_certificates")
            ),
            minimum_experience_months=_experience_months(
                row.get("minimum_experience_months"), row_number
            ),
            job_category=_category_tag(row.get("job_category"), row_number),
            district=_required_text(row.get("district"), "district", row_number),
            status=normalize_text(str(row.get("status") or "")),
        )
    except ValidationError as exc:
        raise JobCatalogMappingError(
            f"Row {row_number}: invalid job data: {exc}"
        ) from exc


def load_job_catalog_workbook(
    workbook_path: str | Path,
    *,
    sheet_name: str = "Jobs",
) -> JobCatalogImportResult:
    """Load, validate and canonicalize all job rows from a workbook."""

    path = Path(workbook_path)
    if not path.is_file():
        raise JobCatalogMappingError(f"Workbook not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise JobCatalogMappingError(
                f"Worksheet '{sheet_name}' not found. Available sheets: "
                + ", ".join(workbook.sheetnames)
            )
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise JobCatalogMappingError("Job workbook is empty") from exc
        headers = [_clean_header(value) for value in raw_headers]
        if len(set(headers)) != len(headers):
            raise JobCatalogMappingError(
                "Job workbook contains duplicate normalized headers"
            )
        missing = sorted(REQUIRED_COLUMNS - set(headers))
        if missing:
            raise JobCatalogMappingError(
                "Missing required job-catalog columns: " + ", ".join(missing)
            )

        jobs: list[Job] = []
        job_ids: set[str] = set()
        mappings: dict[str, str] = {}
        source_rows = 0
        for row_number, values in enumerate(rows, start=2):
            if not any(value is not None and str(value).strip() for value in values):
                continue
            source_rows += 1
            job = map_job_catalog_row(
                dict(zip(headers, values)),
                row_number=row_number,
                skill_mappings=mappings,
            )
            if job.job_id in job_ids:
                raise JobCatalogMappingError(
                    f"Row {row_number}: duplicate job_id '{job.job_id}'"
                )
            job_ids.add(job.job_id)
            jobs.append(job)
        return JobCatalogImportResult(
            jobs=jobs,
            source_rows=source_rows,
            skill_mappings=dict(sorted(mappings.items())),
        )
    finally:
        workbook.close()


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Validate and canonicalize a RIHAI SETU job workbook."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet", default="Jobs")
    parser.add_argument("--output", type=Path)
    args = parser.parse_args(argv)
    try:
        result = load_job_catalog_workbook(args.workbook, sheet_name=args.sheet)
    except JobCatalogMappingError as exc:
        print(f"Job dataset validation failed: {exc}", file=sys.stderr)
        return 2
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(
            json.dumps(
                [job.model_dump(mode="json") for job in result.jobs],
                indent=2,
                ensure_ascii=False,
            )
            + "\n",
            encoding="utf-8",
        )
    print(
        json.dumps(
            {
                **result.summary(),
                "output": str(args.output.resolve()) if args.output else None,
            },
            indent=2,
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
